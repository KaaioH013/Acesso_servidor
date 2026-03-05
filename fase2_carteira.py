"""
fase2_carteira.py — Painel de Carteira em Tempo Real (Fase 2 do Roadmap)

Gera Excel com 5 abas:
  1. Carteira Geral    — todos os itens em aberto com semáforo de prazo
  2. Aguardando NF     — itens STATUS='L' prontos mas sem NF emitida (+ dias parado)
  3. Em Produção       — itens com OP vinculada + status da OP + semáforo
  4. Estoque Disp.     — itens STATUS='V' cruzados com saldo em MT_ESTOQUE
  5. KPIs Resumo       — totais e destaques

Uso:
    python fase2_carteira.py
"""

import sys
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, "src")
from conexao import get_engine

# ── Filtros padrão (peças — igual ao relatorios_vendas.py) ──────────────────
TPVENDA_EXCLUIR = (7, 21, 5, 12, 24, 11, 26, 6, 15, 16, 8, 19, 9, 17, 53, 18, 65, 23)
TPVENDA_STR = ",".join(str(x) for x in TPVENDA_EXCLUIR)

FILTROS_PECAS = f"""
    AND p.STATUS <> 'C'
    AND i.STATUS NOT IN ('C','F')
    AND i.TPVENDA NOT IN ({TPVENDA_STR})
    AND i.MATERIAL NOT LIKE '8%'
    AND i.FLAGSUB <> 'S'
    AND p.CODIGO NOT IN (
        SELECT p2.CODIGO FROM VE_PEDIDO p2
        JOIN FN_FORNECEDORES f2 ON f2.CODIGO = p2.CLIENTE
        WHERE f2.UF = 'EX'
    )
"""

HOJE = date.today()
OUTPUT = Path("exports")
OUTPUT.mkdir(exist_ok=True)


# ── Semáforo de prazo ────────────────────────────────────────────────────────
def semaforo(dt_prazo) -> str:
    if pd.isna(dt_prazo):
        return "⚪ Sem prazo"
    dias = (pd.Timestamp(dt_prazo).date() - HOJE).days
    if dias < 0:
        return "🔴 Atrasado"
    if dias <= 7:
        return "🟡 Urgente"
    return "🟢 No prazo"


def dias_restantes(dt_prazo) -> int | None:
    if pd.isna(dt_prazo):
        return None
    return (pd.Timestamp(dt_prazo).date() - HOJE).days


# ── 1. Carteira Geral ────────────────────────────────────────────────────────
def query_carteira_geral(engine) -> pd.DataFrame:
    sql = f"""
        SELECT
            p.CODIGO            AS Pedido,
            p.NUMINTERNO        AS Nr_Interno,
            f.RAZAO             AS Cliente,
            f.CIDADE            AS Cidade,
            f.UF                AS UF,
            v.RAZAO             AS Vendedor,
            i.SEQ               AS Seq,
            i.MATERIAL          AS Material,
            i.DESCRICAO         AS Descricao,
            tv.DESCRICAO        AS Tipo_Venda,
            i.STATUS            AS Status_Item,
            i.QTDE              AS Qtde,
            i.QTDEFAT           AS Qtde_Fat,
            CASE i.STATUS
                WHEN 'L' THEN i.QTDE
                WHEN 'V' THEN i.QTDE
                WHEN 'A' THEN i.QTDE
                ELSE (i.QTDE - i.QTDEFAT)
            END                 AS Qtde_Saldo,
            i.VLRUNITARIO       AS Vlr_Unit,
            CASE i.STATUS
                WHEN 'L' THEN i.VLRTOTAL
                WHEN 'V' THEN i.VLRTOTAL
                WHEN 'A' THEN i.VLRTOTAL
                ELSE (i.QTDE - i.QTDEFAT) * i.VLRUNITARIO
            END                 AS Vlr_Saldo,
            p.DTPEDIDO          AS Dt_Pedido,
            i.DTPRAZO           AS Dt_Prazo,
            i.DTPRAZOORI        AS Dt_Prazo_Orig,
            i.DTPRAZOCLI        AS Dt_Prazo_Cliente,
            DATEDIFF(DAY, i.DTPRAZO, GETDATE()) AS Dias_Atraso,
            i.DTALTERAFAT       AS Dt_Alt_Fat,
            p.PEDIDOCLI         AS Ped_Cliente
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p         ON p.CODIGO  = i.PEDIDO
        JOIN FN_FORNECEDORES f   ON f.CODIGO  = p.CLIENTE
        JOIN FN_VENDEDORES v     ON v.CODIGO  = p.VENDEDOR
        JOIN VE_TPVENDA tv       ON tv.CODIGO = i.TPVENDA
        WHERE 1=1
        {FILTROS_PECAS}
        ORDER BY i.DTPRAZO ASC, p.DTPEDIDO ASC
    """
    df = pd.read_sql(sql, engine)
    df["Semaforo"] = df["Dt_Prazo"].apply(semaforo)
    df["Dias_p_Prazo"] = df["Dt_Prazo"].apply(dias_restantes)
    # Reordenar colunas com semáforo na frente
    cols = ["Semaforo", "Dias_p_Prazo"] + [c for c in df.columns if c not in ("Semaforo", "Dias_p_Prazo")]
    return df[cols]


# ── 2. Aguardando NF (Status L) ──────────────────────────────────────────────
def query_aguardando_nf(engine) -> pd.DataFrame:
    sql = f"""
        SELECT
            p.CODIGO            AS Pedido,
            p.NUMINTERNO        AS Nr_Interno,
            f.RAZAO             AS Cliente,
            f.CIDADE            AS Cidade,
            v.RAZAO             AS Vendedor,
            i.SEQ               AS Seq,
            i.MATERIAL          AS Material,
            i.DESCRICAO         AS Descricao,
            i.QTDE              AS Qtde,
            i.VLRTOTAL          AS Vlr_Total,
            i.DTALTERAFAT       AS Dt_Liberado,
            DATEDIFF(DAY, i.DTALTERAFAT, GETDATE()) AS Dias_Aguardando,
            i.DTPRAZO           AS Dt_Prazo,
            p.PEDIDOCLI         AS Ped_Cliente
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p         ON p.CODIGO  = i.PEDIDO
        JOIN FN_FORNECEDORES f   ON f.CODIGO  = p.CLIENTE
        JOIN FN_VENDEDORES v     ON v.CODIGO  = p.VENDEDOR
        WHERE i.STATUS = 'L'
          AND p.STATUS <> 'C'
          AND i.FLAGSUB <> 'S'
          AND i.MATERIAL NOT LIKE '8%'
          AND i.TPVENDA NOT IN ({TPVENDA_STR})
          AND p.CODIGO NOT IN (
              SELECT p2.CODIGO FROM VE_PEDIDO p2
              JOIN FN_FORNECEDORES f2 ON f2.CODIGO = p2.CLIENTE
              WHERE f2.UF = 'EX'
          )
        ORDER BY i.DTALTERAFAT ASC
    """
    df = pd.read_sql(sql, engine)
    # Semáforo de urgência por dias aguardando NF
    def urgencia_nf(dias) -> str:
        if pd.isna(dias):
            return "⚪ -"
        d = int(dias)
        if d >= 10:
            return "🔴 Crítico (10+ dias)"
        if d >= 5:
            return "🟡 Atenção (5-9 dias)"
        return "🟢 Recente (< 5 dias)"
    df["Urgencia"] = df["Dias_Aguardando"].apply(urgencia_nf)
    cols = ["Urgencia"] + [c for c in df.columns if c != "Urgencia"]
    return df[cols]


# ── 3. Em Produção (PR_OP) ───────────────────────────────────────────────────
def query_em_producao(engine) -> pd.DataFrame:
    sql = f"""
        SELECT
            p.CODIGO            AS Pedido,
            p.NUMINTERNO        AS Nr_Interno,
            f.RAZAO             AS Cliente,
            v.RAZAO             AS Vendedor,
            i.SEQ               AS Seq,
            i.MATERIAL          AS Material,
            i.DESCRICAO         AS Descricao,
            i.STATUS            AS Status_Item,
            i.QTDE              AS Qtde_Pedida,
            i.QTDEFAT           AS Qtde_Fat,
            i.VLRTOTAL          AS Vlr_Total,
            p.DTPEDIDO          AS Dt_Pedido,
            i.DTPRAZO           AS Dt_Prazo_Pedido,
            DATEDIFF(DAY, i.DTPRAZO, GETDATE()) AS Dias_Atraso_Ped,
            o.NROOP             AS Nr_OP,
            o.OPSTATUS          AS Status_OP_Cod,
            CASE o.OPSTATUS
                WHEN 1 THEN '🔵 Aguardando Liberação'
                WHEN 2 THEN '🟡 Em Produção'
                WHEN 3 THEN '🟢 Encerrada'
                WHEN 4 THEN '🔴 Cancelada'
                ELSE '⚪ Desconhecido'
            END                 AS Status_OP,
            o.QTDE              AS Qtde_OP,
            o.QTDEPRODUZIDA     AS Qtde_Produzida,
            o.DTCADASTRO        AS Dt_Abertura_OP,
            o.DTPRAZO           AS Dt_Prazo_OP,
            o.DTEFETIVA         AS Dt_Encerramento_OP,
            DATEDIFF(DAY, o.DTPRAZO, GETDATE()) AS Dias_Atraso_OP
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p         ON p.CODIGO  = i.PEDIDO
        JOIN FN_FORNECEDORES f   ON f.CODIGO  = p.CLIENTE
        JOIN FN_VENDEDORES v     ON v.CODIGO  = p.VENDEDOR
        JOIN PR_OP o             ON o.PEDIDO  = p.CODIGO
        WHERE i.STATUS NOT IN ('C','F')
          AND p.STATUS <> 'C'
          AND o.OPSTATUS IN (1, 2)
          AND i.FLAGSUB <> 'S'
          AND i.MATERIAL NOT LIKE '8%'
          AND i.TPVENDA NOT IN ({TPVENDA_STR})
          AND p.CODIGO NOT IN (
              SELECT p2.CODIGO FROM VE_PEDIDO p2
              JOIN FN_FORNECEDORES f2 ON f2.CODIGO = p2.CLIENTE
              WHERE f2.UF = 'EX'
          )
        ORDER BY o.DTPRAZO ASC
    """
    df = pd.read_sql(sql, engine)
    df["Semaforo_OP"] = df["Dt_Prazo_OP"].apply(semaforo)
    cols = ["Semaforo_OP"] + [c for c in df.columns if c != "Semaforo_OP"]
    return df[cols]


# ── 4. Disponibilidade de Estoque ────────────────────────────────────────────
def query_estoque_disponivel(engine) -> pd.DataFrame:
    """
    Itens com STATUS='V' (nada entregue ainda) cruzados com MT_ESTOQUE.
    Mostra se há estoque suficiente para atender o pedido.
    """
    sql = f"""
        SELECT
            p.CODIGO            AS Pedido,
            p.NUMINTERNO        AS Nr_Interno,
            f.RAZAO             AS Cliente,
            v.RAZAO             AS Vendedor,
            i.SEQ               AS Seq,
            i.MATERIAL          AS Material,
            i.DESCRICAO         AS Descricao,
            i.QTDE              AS Qtde_Pedida,
            i.VLRTOTAL          AS Vlr_Total,
            p.DTPEDIDO          AS Dt_Pedido,
            i.DTPRAZO           AS Dt_Prazo,
            ISNULL(SUM(e.QTDEREAL), 0) AS Estoque_Disponivel,
            CASE
                WHEN ISNULL(SUM(e.QTDEREAL), 0) >= i.QTDE THEN '✅ Atende'
                WHEN ISNULL(SUM(e.QTDEREAL), 0) > 0      THEN '⚠️ Parcial'
                ELSE '❌ Sem estoque'
            END                 AS Situacao_Estoque
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p         ON p.CODIGO  = i.PEDIDO
        JOIN FN_FORNECEDORES f   ON f.CODIGO  = p.CLIENTE
        JOIN FN_VENDEDORES v     ON v.CODIGO  = p.VENDEDOR
        LEFT JOIN MT_ESTOQUE e   ON e.MATERIAL = i.MATERIAL
                                AND e.TPESTOQUE = 'AL'
        WHERE i.STATUS = 'V'
          AND p.STATUS <> 'C'
          AND i.FLAGSUB <> 'S'
          AND i.MATERIAL NOT LIKE '8%'
          AND i.TPVENDA NOT IN ({TPVENDA_STR})
          AND p.CODIGO NOT IN (
              SELECT p2.CODIGO FROM VE_PEDIDO p2
              JOIN FN_FORNECEDORES f2 ON f2.CODIGO = p2.CLIENTE
              WHERE f2.UF = 'EX'
          )
        GROUP BY
            p.CODIGO, p.NUMINTERNO, f.RAZAO, v.RAZAO,
            i.SEQ, i.MATERIAL, i.DESCRICAO, i.QTDE,
            i.VLRTOTAL, p.DTPEDIDO, i.DTPRAZO
        ORDER BY i.DTPRAZO ASC
    """
    df = pd.read_sql(sql, engine)
    df["Semaforo"] = df["Dt_Prazo"].apply(semaforo)
    cols = ["Situacao_Estoque", "Semaforo"] + [c for c in df.columns if c not in ("Situacao_Estoque", "Semaforo")]
    return df[cols]


# ── 5. KPIs Resumo ───────────────────────────────────────────────────────────
def montar_resumo(df_cart: pd.DataFrame, df_nf: pd.DataFrame,
                  df_op: pd.DataFrame) -> pd.DataFrame:
    hoje_str = HOJE.strftime("%d/%m/%Y")
    total_vlr = df_cart["Vlr_Saldo"].sum()
    atrasados = df_cart[df_cart["Semaforo"] == "🔴 Atrasado"]
    urgentes  = df_cart[df_cart["Semaforo"] == "🟡 Urgente"]
    no_prazo  = df_cart[df_cart["Semaforo"] == "🟢 No prazo"]

    linhas = [
        ("Data de referência",         hoje_str),
        ("─── CARTEIRA GERAL ───",    ""),
        ("Total itens em aberto",       f"{len(df_cart):,}"),
        ("Valor total carteira",        f"R$ {total_vlr:,.2f}".replace(",","X").replace(".",",").replace("X",".")),
        ("Itens 🔴 Atrasados",          f"{len(atrasados):,}  |  R$ {atrasados['Vlr_Saldo'].sum():,.0f}".replace(",",".")),
        ("Itens 🟡 Urgentes (≤7d)",     f"{len(urgentes):,}  |  R$ {urgentes['Vlr_Saldo'].sum():,.0f}".replace(",",".")),
        ("Itens 🟢 No prazo",           f"{len(no_prazo):,}  |  R$ {no_prazo['Vlr_Saldo'].sum():,.0f}".replace(",",".")),
        ("─── AGUARDANDO NF ───",      ""),
        ("Itens esperando NF",          f"{len(df_nf):,}"),
        ("Valor esperando NF",          f"R$ {df_nf['Vlr_Total'].sum():,.2f}".replace(",","X").replace(".",",").replace("X",".")),
        ("Críticos (≥10 dias sem NF)",  f"{(df_nf['Dias_Aguardando'] >= 10).sum():,}"),
        ("─── PRODUÇÃO ───",           ""),
        ("OPs em aberto/produção",      f"{len(df_op):,}"),
        ("OPs com prazo atrasado",      f"{(df_op['Dias_Atraso_OP'] > 0).sum():,}"),
    ]

    # Clientes com maior saldo na carteira
    top_cli = (df_cart.groupby("Cliente")["Vlr_Saldo"]
               .sum().sort_values(ascending=False).head(5))

    linhas.append(("─── TOP 5 CLIENTES EM CARTEIRA ───", ""))
    for cli, vlr in top_cli.items():
        linhas.append((f"  {cli[:40]}", f"R$ {vlr:,.0f}".replace(",",".")))

    return pd.DataFrame(linhas, columns=["Indicador", "Valor"])


# ── Excel helpers ─────────────────────────────────────────────────────────────
CORES_SEMAFORO = {
    "🔴": "FFC7CE",
    "🟡": "FFEB9C",
    "🟢": "C6EFCE",
    "⚪": "E0E0E0",
    "🔵": "BDD7EE",
    "✅": "C6EFCE",
    "⚠️": "FFEB9C",
    "❌": "FFC7CE",
}


def cor_linha(valor: str) -> str | None:
    for emoji, cor in CORES_SEMAFORO.items():
        if str(valor).startswith(emoji):
            return cor
    return None


def escrever_aba(writer, nome_aba: str, df: pd.DataFrame,
                 col_semaforo: str | None = None):
    """Escreve DataFrame com cabeçalho azul e coloração de semáforo por linha."""
    nome = nome_aba[:31]
    df.to_excel(writer, sheet_name=nome, index=False)
    ws = writer.sheets[nome]

    # Cabeçalho
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    # Colorir linhas pelo valor da coluna semáforo
    if col_semaforo and col_semaforo in df.columns:
        col_idx = list(df.columns).index(col_semaforo)  # 0-based
        for row_idx in range(2, len(df) + 2):
            cell_val = str(ws.cell(row=row_idx, column=col_idx + 1).value or "")
            hex_cor = cor_linha(cell_val)
            if hex_cor:
                fill = PatternFill("solid", fgColor=hex_cor)
                for col in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

    # Ajustar larguras
    for i, col in enumerate(df.columns, 1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 50)

    # Congelar primeira linha
    ws.freeze_panes = "A2"

    # Auto-filtro
    ws.auto_filter.ref = ws.dimensions


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    engine = get_engine()
    hoje_fmt = HOJE.strftime("%d/%m/%Y")
    print(f"Painel de Carteira — {hoje_fmt}")

    print("  [1/4] Carteira geral (todos os itens abertos)...")
    df_cart = query_carteira_geral(engine)

    print("  [2/4] Itens aguardando emissão de NF (Status L)...")
    df_nf = query_aguardando_nf(engine)

    print("  [3/4] Itens com Ordem de Produção em aberto (PR_OP)...")
    df_op = query_em_producao(engine)

    print("  [4/4] Disponibilidade de estoque (itens Status V)...")
    df_est = query_estoque_disponivel(engine)

    df_resumo = montar_resumo(df_cart, df_nf, df_op)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo = OUTPUT / f"carteira_{HOJE.strftime('%d-%m-%Y')}_{ts}.xlsx"

    print(f"\nGerando Excel: {arquivo}")
    with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
        escrever_aba(writer, "KPIs Resumo",       df_resumo)
        escrever_aba(writer, "Carteira Geral",    df_cart,  col_semaforo="Semaforo")
        escrever_aba(writer, "Aguardando NF",     df_nf,    col_semaforo="Urgencia")
        escrever_aba(writer, "Em Produção (OP)",  df_op,    col_semaforo="Semaforo_OP")
        escrever_aba(writer, "Estoque Dispon.",   df_est,   col_semaforo="Situacao_Estoque")

    print(f"\n✅ Concluído! Arquivo: {arquivo.resolve()}")

    # ── Resumo no console ──
    print(f"\n{'='*55}")
    print(f"  Data          : {hoje_fmt}")
    print(f"  Itens abertos : {len(df_cart):>5,}")
    print(f"  Valor carteira: R$ {df_cart['Vlr_Saldo'].sum():>12,.2f}".replace(",","X").replace(".",",").replace("X","."))
    print(f"  🔴 Atrasados  : {(df_cart['Semaforo']=='🔴 Atrasado').sum():>5,}")
    print(f"  🟡 Urgentes   : {(df_cart['Semaforo']=='🟡 Urgente').sum():>5,}")
    print(f"  🟢 No prazo   : {(df_cart['Semaforo']=='🟢 No prazo').sum():>5,}")
    print(f"  Aguard. NF    : {len(df_nf):>5,} itens  |  R$ {df_nf['Vlr_Total'].sum():,.0f}".replace(",","."))
    print(f"  OPs abertas   : {len(df_op):>5,}")
    print("="*55)


if __name__ == "__main__":
    main()
