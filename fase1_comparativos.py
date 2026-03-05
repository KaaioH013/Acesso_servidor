"""
fase1_comparativos.py — Análises Estratégicas de Vendas (Fase 1 do Roadmap)

Gera Excel com 6 abas:
  1. Evolução Mensal   — últimos 24 meses (YoY implícito)
  2. Curva ABC Clientes — classificação A/B/C por faturamento no ano de referência
  3. Curva ABC Materiais — classificação A/B/C por faturamento no ano de referência
  4. Conversão Orçamentos — taxa de conversão mensal dos últimos 24 meses
  5. Retenção Clientes  — clientes novos / retidos / perdidos (churn) YoY
  6. Ranking Vendedores — comparativo YoY por vendedor

Uso:
    python fase1_comparativos.py              # usa ano atual
    python fase1_comparativos.py --ano 2025   # ano específico de referência
"""

import sys
import argparse
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from sqlalchemy import text

sys.path.insert(0, "src")
from conexao import get_engine

# ── Constantes de filtro (mesmas do relatorios_vendas.py) ────────────────────
TPVENDA_EXCLUIR = (7, 21, 5, 12, 24, 11, 26, 6, 15, 16, 8, 19, 9, 17, 53, 18, 65, 23)
TPVENDA_STR = ",".join(str(x) for x in TPVENDA_EXCLUIR)

FILTROS_BASE = f"""
    AND p.STATUS <> 'C'
    AND i.STATUS <> 'C'
    AND i.TPVENDA NOT IN ({TPVENDA_STR})
    AND i.MATERIAL NOT LIKE '8%'
    AND i.FLAGSUB <> 'S'
    AND p.CODIGO NOT IN (
        SELECT p2.CODIGO FROM VE_PEDIDO p2
        JOIN FN_FORNECEDORES f2 ON f2.CODIGO = p2.CLIENTE
        WHERE f2.UF = 'EX'
    )
"""

OUTPUT = Path("exports")
OUTPUT.mkdir(exist_ok=True)


def parse_args():
    parser = argparse.ArgumentParser(description="Análises estratégicas Fase 1")
    parser.add_argument(
        "--ano",
        type=int,
        default=datetime.now().year,
        help="Ano de referência (padrão: ano atual)",
    )
    return parser.parse_args()


# ── 1. Evolução Mensal (últimos 24 meses) ────────────────────────────────────
def query_evolucao_mensal(engine, data_ini: str, data_fim: str) -> pd.DataFrame:
    sql = f"""
        SELECT
            YEAR(p.DTPEDIDO)  AS Ano,
            MONTH(p.DTPEDIDO) AS Mes,
            COUNT(DISTINCT p.CODIGO) AS Pedidos,
            COUNT(i.CODIGO)          AS Itens,
            SUM(i.VLRTOTAL)          AS Faturamento,
            SUM(i.QTDE)              AS Qtde_Total
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p ON p.CODIGO = i.PEDIDO
        WHERE p.DTPEDIDO BETWEEN '{data_ini}' AND '{data_fim}'
        {FILTROS_BASE}
        GROUP BY YEAR(p.DTPEDIDO), MONTH(p.DTPEDIDO)
        ORDER BY Ano, Mes
    """
    df = pd.read_sql(sql, engine)
    df["Periodo"] = df.apply(
        lambda r: f"{int(r.Mes):02d}/{int(r.Ano)}", axis=1
    )
    # Coluna YoY_pct será calculada em Python após pivot
    return df


def calcular_yoy(df: pd.DataFrame) -> pd.DataFrame:
    """Adiciona coluna de variação YoY em %. Requer colunas Ano, Mes, Faturamento."""
    df = df.sort_values(["Ano", "Mes"]).copy()
    prev = df.copy()
    prev["Ano"] = prev["Ano"] + 1
    merged = df.merge(
        prev[["Ano", "Mes", "Faturamento", "Pedidos"]],
        on=["Ano", "Mes"],
        suffixes=("", "_ano_ant"),
        how="left",
    )
    merged["YoY_Fat_pct"] = (
        (merged["Faturamento"] - merged["Faturamento_ano_ant"])
        / merged["Faturamento_ano_ant"]
        * 100
    ).round(1)
    merged["YoY_Ped_pct"] = (
        (merged["Pedidos"] - merged["Pedidos_ano_ant"])
        / merged["Pedidos_ano_ant"]
        * 100
    ).round(1)
    return merged


# ── 2 & 3. Curva ABC ─────────────────────────────────────────────────────────
def query_abc_clientes(engine, ano: int) -> pd.DataFrame:
    sql = f"""
        SELECT
            f.CODIGO        AS Cod_Cliente,
            f.RAZAO         AS Cliente,
            f.CIDADE        AS Cidade,
            f.UF            AS UF,
            COUNT(DISTINCT p.CODIGO) AS Pedidos,
            SUM(i.QTDE)              AS Qtde,
            SUM(i.VLRTOTAL)          AS Faturamento
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p      ON p.CODIGO   = i.PEDIDO
        JOIN FN_FORNECEDORES f ON f.CODIGO  = p.CLIENTE
        WHERE YEAR(p.DTPEDIDO) = {ano}
        {FILTROS_BASE}
        GROUP BY f.CODIGO, f.RAZAO, f.CIDADE, f.UF
        ORDER BY Faturamento DESC
    """
    df = pd.read_sql(sql, engine)
    df = _classificar_abc(df, "Faturamento")
    df["Part_pct"] = (df["Faturamento"] / df["Faturamento"].sum() * 100).round(2)
    df["Fat_Acum_pct"] = df["Part_pct"].cumsum().round(2)
    return df


def query_abc_materiais(engine, ano: int) -> pd.DataFrame:
    sql = f"""
        SELECT
            i.MATERIAL      AS Material,
            MAX(i.DESCRICAO) AS Descricao,
            COUNT(DISTINCT p.CODIGO) AS Pedidos,
            SUM(i.QTDE)              AS Qtde,
            SUM(i.VLRTOTAL)          AS Faturamento
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p ON p.CODIGO = i.PEDIDO
        WHERE YEAR(p.DTPEDIDO) = {ano}
        {FILTROS_BASE}
        GROUP BY i.MATERIAL
        ORDER BY Faturamento DESC
    """
    df = pd.read_sql(sql, engine)
    df = _classificar_abc(df, "Faturamento")
    df["Part_pct"] = (df["Faturamento"] / df["Faturamento"].sum() * 100).round(2)
    df["Fat_Acum_pct"] = df["Part_pct"].cumsum().round(2)
    return df


def _classificar_abc(df: pd.DataFrame, col_valor: str) -> pd.DataFrame:
    """Classifica linhas em A (80%), B (95%), C (100%) por valor acumulado."""
    total = df[col_valor].sum()
    acum = 0.0
    classes = []
    for v in df[col_valor]:
        acum += v
        pct = acum / total * 100
        if pct <= 80:
            classes.append("A")
        elif pct <= 95:
            classes.append("B")
        else:
            classes.append("C")
    df = df.copy()
    df["Curva"] = classes
    return df


# ── 4. Conversão de Orçamentos ───────────────────────────────────────────────
def query_conversao_orcamentos(engine, data_ini: str, data_fim: str) -> pd.DataFrame:
    sql = f"""
        SELECT
            YEAR(o.DTCADASTRO)  AS Ano,
            MONTH(o.DTCADASTRO) AS Mes,
            COUNT(*)                                                    AS Total,
            SUM(CASE WHEN o.STATUS = 'A' THEN 1 ELSE 0 END)            AS Aprovados,
            SUM(CASE WHEN o.STATUS = 'P' THEN 1 ELSE 0 END)            AS Perdidos,
            SUM(CASE WHEN o.STATUS NOT IN ('A','P') THEN 1 ELSE 0 END) AS Em_Aberto,
            SUM(CASE WHEN o.PEDORIGEM IS NOT NULL THEN 1 ELSE 0 END)   AS Com_Pedido,
            ISNULL(SUM(o.VLRORCADO),  0)   AS Vlr_Orcado,
            ISNULL(SUM(o.VLREFETIVO), 0)   AS Vlr_Convertido
        FROM VE_ORCAMENTOS o
        WHERE o.DTCADASTRO BETWEEN '{data_ini}' AND '{data_fim}'
        GROUP BY YEAR(o.DTCADASTRO), MONTH(o.DTCADASTRO)
        ORDER BY Ano, Mes
    """
    df = pd.read_sql(sql, engine)
    df["Periodo"] = df.apply(lambda r: f"{int(r.Mes):02d}/{int(r.Ano)}", axis=1)
    df["Taxa_Conv_pct"] = (
        df["Aprovados"] / df["Total"].replace(0, pd.NA) * 100
    ).round(1)
    df["Taxa_Vlr_pct"] = (
        df["Vlr_Convertido"] / df["Vlr_Orcado"].replace(0, pd.NA) * 100
    ).round(1)
    return df


# ── 5. Retenção / Churn de Clientes ─────────────────────────────────────────
def query_retencao(engine, ano: int) -> dict:
    """
    Compara clientes do ano de referência vs o anterior.
    Retorna dict com DataFrames: resumo, novos, retidos, perdidos.
    """
    ano_ant = ano - 1

    def clientes_do_ano(a: int) -> set:
        sql = f"""
            SELECT DISTINCT p.CLIENTE
            FROM VE_PEDIDOITENS i
            JOIN VE_PEDIDO p ON p.CODIGO = i.PEDIDO
            WHERE YEAR(p.DTPEDIDO) = {a}
            {FILTROS_BASE}
        """
        return set(pd.read_sql(sql, engine)["CLIENTE"].tolist())

    def detalhe_clientes(codigos: list, ref_ano: int) -> pd.DataFrame:
        if not codigos:
            return pd.DataFrame()
        cs = ",".join(str(c) for c in codigos)
        sql = f"""
            SELECT
                f.CODIGO        AS Cod_Cliente,
                f.RAZAO         AS Cliente,
                f.CIDADE        AS Cidade,
                f.UF            AS UF,
                SUM(i.VLRTOTAL) AS Faturamento_{ref_ano},
                COUNT(DISTINCT p.CODIGO) AS Pedidos_{ref_ano}
            FROM VE_PEDIDOITENS i
            JOIN VE_PEDIDO p      ON p.CODIGO  = i.PEDIDO
            JOIN FN_FORNECEDORES f ON f.CODIGO  = p.CLIENTE
            WHERE YEAR(p.DTPEDIDO) = {ref_ano}
              AND p.CLIENTE IN ({cs})
            {FILTROS_BASE}
            GROUP BY f.CODIGO, f.RAZAO, f.CIDADE, f.UF
            ORDER BY Faturamento_{ref_ano} DESC
        """
        return pd.read_sql(sql, engine)

    cli_atual = clientes_do_ano(ano)
    cli_ant = clientes_do_ano(ano_ant)

    novos = sorted(cli_atual - cli_ant)        # compraram em `ano` mas não em `ano_ant`
    retidos = sorted(cli_atual & cli_ant)      # compraram nos dois anos
    perdidos = sorted(cli_ant - cli_atual)     # compraram em `ano_ant` mas não em `ano`

    df_novos = detalhe_clientes(novos, ano)
    df_retidos = detalhe_clientes(retidos, ano)
    df_perdidos = detalhe_clientes(perdidos, ano_ant)

    resumo = pd.DataFrame([{
        "Indicador": f"Clientes em {ano}",
        "Qtd": len(cli_atual),
    }, {
        "Indicador": f"Clientes em {ano_ant}",
        "Qtd": len(cli_ant),
    }, {
        "Indicador": "Novos (só em " + str(ano) + ")",
        "Qtd": len(novos),
    }, {
        "Indicador": "Retidos (ambos os anos)",
        "Qtd": len(retidos),
    }, {
        "Indicador": "Perdidos / Churn (só em " + str(ano_ant) + ")",
        "Qtd": len(perdidos),
    }, {
        "Indicador": "Taxa de Retenção",
        "Qtd": round(len(retidos) / max(len(cli_ant), 1) * 100, 1),
    }, {
        "Indicador": "Taxa de Churn",
        "Qtd": round(len(perdidos) / max(len(cli_ant), 1) * 100, 1),
    }])

    return {
        "resumo": resumo,
        "novos": df_novos,
        "retidos": df_retidos,
        "perdidos": df_perdidos,
    }


# ── 6. Ranking Vendedores YoY ────────────────────────────────────────────────
def query_ranking_vendedores(engine, ano: int) -> pd.DataFrame:
    anos = (ano - 1, ano)
    sql = f"""
        SELECT
            v.CODIGO       AS Cod_Vend,
            v.RAZAO        AS Vendedor,
            YEAR(p.DTPEDIDO) AS Ano,
            COUNT(DISTINCT p.CODIGO) AS Pedidos,
            COUNT(i.CODIGO)          AS Itens,
            SUM(i.VLRTOTAL)          AS Faturamento
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p       ON p.CODIGO  = i.PEDIDO
        JOIN FN_VENDEDORES v   ON v.CODIGO  = p.VENDEDOR
        WHERE YEAR(p.DTPEDIDO) IN ({anos[0]}, {anos[1]})
        {FILTROS_BASE}
        GROUP BY v.CODIGO, v.RAZAO, YEAR(p.DTPEDIDO)
        ORDER BY Faturamento DESC
    """
    df = pd.read_sql(sql, engine)

    # Pivot: uma linha por vendedor, colunas por ano
    fat = df.pivot_table(
        index=["Cod_Vend", "Vendedor"],
        columns="Ano",
        values="Faturamento",
        aggfunc="sum",
    ).reset_index()
    fat.columns.name = None
    fat.columns = [
        str(c) if isinstance(c, int) else c for c in fat.columns
    ]

    ped = df.pivot_table(
        index=["Cod_Vend", "Vendedor"],
        columns="Ano",
        values="Pedidos",
        aggfunc="sum",
    ).reset_index()
    ped.columns.name = None
    ped.columns = [
        str(c) if isinstance(c, int) else c for c in ped.columns
    ]
    ped = ped.rename(
        columns={str(anos[0]): f"Ped_{anos[0]}", str(anos[1]): f"Ped_{anos[1]}"}
    )

    out = fat.merge(ped[["Cod_Vend", f"Ped_{anos[0]}", f"Ped_{anos[1]}"]], on="Cod_Vend")

    a0_col, a1_col = str(anos[0]), str(anos[1])
    if a0_col in out.columns and a1_col in out.columns:
        out["YoY_pct"] = (
            (out[a1_col] - out[a0_col]) / out[a0_col].replace(0, pd.NA) * 100
        ).round(1)
        out = out.sort_values(a1_col, ascending=False)

    return out


# ── Exportar para Excel ──────────────────────────────────────────────────────
def ajustar_colunas(writer, aba: str, df: pd.DataFrame):
    """Auto-ajusta larguras das colunas (openpyxl)."""
    from openpyxl.utils import get_column_letter
    ws = writer.sheets[aba]
    for i, col in enumerate(df.columns, 1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 50)


def colorir_abc(writer, aba: str, df: pd.DataFrame):
    """Colore a coluna 'Curva' com verde/amarelo/vermelho."""
    from openpyxl.styles import PatternFill, Font
    ws = writer.sheets[aba]
    if "Curva" not in df.columns:
        return
    col_idx = list(df.columns).index("Curva") + 1  # 1-based
    fills = {
        "A": PatternFill("solid", fgColor="C6EFCE"),
        "B": PatternFill("solid", fgColor="FFEB9C"),
        "C": PatternFill("solid", fgColor="FFC7CE"),
    }
    for row in range(2, len(df) + 2):  # row 1 = header
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = fills.get(str(cell.value), PatternFill())


def escrever_aba(writer, nome_aba: str, df: pd.DataFrame, colorir=False):
    nome = nome_aba[:31]
    df.to_excel(writer, sheet_name=nome, index=False)
    ajustar_colunas(writer, nome, df)
    if colorir:
        colorir_abc(writer, nome, df)


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    args = parse_args()
    ano = args.ano
    ano_ant = ano - 1

    engine = get_engine()
    print(f"Gerando relatório de análises estratégicas — referência: {ano}")

    # Datas para série de 24 meses
    data_fim = f"{ano}-12-31"
    data_ini = f"{ano_ant}-01-01"

    # 1. Evolução mensal
    print("  [1/6] Evolução mensal (24 meses)...")
    df_evolucao = query_evolucao_mensal(engine, data_ini, data_fim)
    df_evolucao = calcular_yoy(df_evolucao)

    # 2. ABC Clientes
    print("  [2/6] Curva ABC Clientes...")
    df_abc_cli = query_abc_clientes(engine, ano)

    # 3. ABC Materiais
    print("  [3/6] Curva ABC Materiais...")
    df_abc_mat = query_abc_materiais(engine, ano)

    # 4. Conversão de orçamentos
    print("  [4/6] Conversão de orçamentos...")
    df_conv = query_conversao_orcamentos(engine, data_ini, data_fim)

    # 5. Retenção / Churn
    print("  [5/6] Retenção de clientes (pode demorar)...")
    retencao = query_retencao(engine, ano)

    # 6. Ranking vendedores
    print("  [6/6] Ranking vendedores...")
    df_vend = query_ranking_vendedores(engine, ano)

    # ── Exportar ──
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo = OUTPUT / f"fase1_analitico_{ano}_{ts}.xlsx"

    print(f"\nGerando Excel: {arquivo}")
    with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:

        # Aba 1 — Evolução Mensal
        colunas_ev = [
            "Periodo", "Ano", "Mes", "Pedidos", "Itens",
            "Faturamento", "Qtde_Total",
            "Faturamento_ano_ant", "YoY_Fat_pct", "YoY_Ped_pct"
        ]
        df_ev_out = df_evolucao[[c for c in colunas_ev if c in df_evolucao.columns]]
        escrever_aba(writer, "Evolução Mensal", df_ev_out)

        # Aba 2 — ABC Clientes
        escrever_aba(writer, "ABC Clientes", df_abc_cli, colorir=True)

        # Aba 3 — ABC Materiais
        escrever_aba(writer, "ABC Materiais", df_abc_mat, colorir=True)

        # Aba 4 — Conversão Orçamentos
        escrever_aba(writer, "Conversão Orçamentos", df_conv)

        # Aba 5a — Resumo Retenção
        escrever_aba(writer, "Retenção Resumo", retencao["resumo"])

        # Aba 5b — Clientes Novos
        if not retencao["novos"].empty:
            escrever_aba(writer, "Clientes Novos", retencao["novos"])

        # Aba 5c — Clientes Retidos
        if not retencao["retidos"].empty:
            escrever_aba(writer, "Clientes Retidos", retencao["retidos"])

        # Aba 5d — Clientes Perdidos (Churn)
        if not retencao["perdidos"].empty:
            escrever_aba(writer, "Clientes Perdidos", retencao["perdidos"])

        # Aba 6 — Ranking Vendedores
        escrever_aba(writer, "Ranking Vendedores", df_vend)

    print(f"\n✅ Concluído! Arquivo salvo em:\n   {arquivo.resolve()}")

    # ── Resumo no console ──
    print(f"\n{'='*55}")
    print(f"  Ano de referência : {ano}")
    print(f"  Receita bruta {ano} : "
          f"R$ {df_abc_cli['Faturamento'].sum():>14,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    print(f"  Clientes ativos   : {len(df_abc_cli):>5}")
    print(f"  Materiais vendidos: {len(df_abc_mat):>5}")
    print(f"  Clientes Curva A  : {(df_abc_cli['Curva']=='A').sum():>5}")
    print(f"  Clientes Curva B  : {(df_abc_cli['Curva']=='B').sum():>5}")
    print(f"  Clientes Curva C  : {(df_abc_cli['Curva']=='C').sum():>5}")
    print(f"\n  Retenção {ano_ant}→{ano}  :")
    for _, row in retencao["resumo"].iterrows():
        print(f"    {row['Indicador']:<40}: {row['Qtd']}")
    print("="*55)


if __name__ == "__main__":
    main()
