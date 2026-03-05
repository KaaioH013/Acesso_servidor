"""
relatorio_506_excel.py — Evolução do Relatório 506 (Contas a Receber)

Inclui no Excel:
- Vendedor Externo
- Vendedor Interno
- Estado (UF)
- Status_Pagamento: PAGA / NAO PAGA

Uso:
    python relatorio_506_excel.py
    python relatorio_506_excel.py --dt-ini 2026-01-01 --dt-fim 2026-02-27
    python relatorio_506_excel.py --incluir-exterior
"""

import argparse
from datetime import date, datetime
from pathlib import Path
import sys

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, "src")
from conexao import get_engine

OUTPUT = Path("exports")
OUTPUT.mkdir(exist_ok=True)


def parse_args():
    hoje = date.today()
    dt_ini_default = f"{hoje.year}-01-01"
    dt_fim_default = hoje.strftime("%Y-%m-%d")

    p = argparse.ArgumentParser()
    p.add_argument("--dt-ini", type=str, default=dt_ini_default, help="Data inicial emissão NF (YYYY-MM-DD)")
    p.add_argument("--dt-fim", type=str, default=dt_fim_default, help="Data final emissão NF (YYYY-MM-DD)")
    p.add_argument("--incluir-exterior", action="store_true", help="Inclui UF='EX'")
    return p.parse_args()


def query_506(engine, dt_ini: str, dt_fim: str, incluir_exterior: bool) -> pd.DataFrame:
    filtro_ex = "" if incluir_exterior else "AND ISNULL(c.UF,'') <> 'EX'"

    sql = f"""
WITH VendedoresPedido AS (
    SELECT
        pv.PEDIDO,
        MAX(CASE WHEN fv.TIPO = 'E' AND pv.CKPRINCIPAL = 'S' THEN fv.RAZAO END) AS Vend_Ext_Principal,
        MAX(CASE WHEN fv.TIPO = 'E' THEN fv.RAZAO END) AS Vend_Ext,
        MAX(CASE WHEN fv.TIPO = 'I' AND pv.CKPRINCIPAL = 'S' THEN fv.RAZAO END) AS Vend_Int_Principal,
        MAX(CASE WHEN fv.TIPO = 'I' THEN fv.RAZAO END) AS Vend_Int
    FROM VE_PEDIDOVENDEDOR pv
    JOIN FN_VENDEDORES fv ON fv.CODIGO = pv.VENDEDOR
    GROUP BY pv.PEDIDO
),
PedidosNFS AS (
    SELECT DISTINCT
        n.CODIGO AS NFS,
        n.VEPEDIDO AS PEDIDO
    FROM FN_NFS n
    WHERE n.VEPEDIDO IS NOT NULL

    UNION

    SELECT DISTINCT
        x.NFS,
        x.VEPEDIDO AS PEDIDO
    FROM FN_NFSADTPEDIDO x
    WHERE x.NFS IS NOT NULL
      AND x.VEPEDIDO IS NOT NULL

        UNION

        SELECT DISTINCT
                i.NFS,
                pi.PEDIDO
        FROM FN_NFSITENS i
        JOIN VE_PEDIDOITENS pi ON pi.CODIGO = i.PEDIDOITEM
        WHERE i.NFS IS NOT NULL
            AND i.PEDIDOITEM IS NOT NULL
            AND pi.PEDIDO IS NOT NULL
),
VendedoresNFS AS (
    SELECT
        pn.NFS,
        MAX(vp.Vend_Ext_Principal) AS Vend_Ext_Principal,
        MAX(vp.Vend_Ext) AS Vend_Ext,
        MAX(vp.Vend_Int_Principal) AS Vend_Int_Principal,
        MAX(vp.Vend_Int) AS Vend_Int
    FROM PedidosNFS pn
    JOIN VendedoresPedido vp ON vp.PEDIDO = pn.PEDIDO
    GROUP BY pn.NFS
),
StatusNF AS (
    SELECT
        rr.NFS,
        COUNT(*) AS Parcelas_Total,
        SUM(CASE WHEN rr.DTPAGAMENTO IS NOT NULL THEN 1 ELSE 0 END) AS Parcelas_Pagas,
        SUM(CASE WHEN rr.DTPAGAMENTO IS NULL THEN 1 ELSE 0 END) AS Parcelas_Abertas,
        MIN(rr.DTVENCIMENTO) AS Dt_Primeiro_Venc_NF,
        MAX(rr.DTVENCIMENTO) AS Dt_Ultimo_Venc_NF,
        MAX(rr.DTPAGAMENTO) AS Dt_Ultimo_Pagto_NF,
        CASE
            WHEN COUNT(*) = SUM(CASE WHEN rr.DTPAGAMENTO IS NOT NULL THEN 1 ELSE 0 END)
            THEN 'QUITADA'
            ELSE 'PENDENTE'
        END AS Status_NF_Quitada
    FROM FN_RECEBER rr
    WHERE rr.NFS IS NOT NULL
      AND rr.DTCANCELAMENTO IS NULL
    GROUP BY rr.NFS
),
TipoNFS AS (
    SELECT
        i.NFS,
        SUM(CASE WHEN i.MATERIAL LIKE '8%' THEN 1 ELSE 0 END) AS Qtd_Itens_Bomba,
        SUM(CASE WHEN i.MATERIAL IS NOT NULL AND i.MATERIAL NOT LIKE '8%' THEN 1 ELSE 0 END) AS Qtd_Itens_Peca,
        CASE
            WHEN SUM(CASE WHEN i.MATERIAL LIKE '8%' THEN 1 ELSE 0 END) > 0
             AND SUM(CASE WHEN i.MATERIAL IS NOT NULL AND i.MATERIAL NOT LIKE '8%' THEN 1 ELSE 0 END) > 0
            THEN 'AMBAS'
            WHEN SUM(CASE WHEN i.MATERIAL LIKE '8%' THEN 1 ELSE 0 END) > 0
            THEN 'BOMBA'
            WHEN SUM(CASE WHEN i.MATERIAL IS NOT NULL AND i.MATERIAL NOT LIKE '8%' THEN 1 ELSE 0 END) > 0
            THEN 'PECA'
            ELSE 'NAO_CLASSIFICADO'
        END AS Tipo_Produto_NF
    FROM FN_NFSITENS i
    GROUP BY i.NFS
)
SELECT
    r.FILIAL AS Fl,
    r.CODIGO AS Receber,
    r.NFS,
    r.NOTASEQ,
    r.NRODOCUMENTO AS Nro_Documento,
    n.TPDOCUMENTO AS Tp_Documento,
    n.DTEMISSAO AS Dt_Emissao_NF,
    r.DTEMISSAO AS Dt_Emissao_Titulo,
    r.DTVENCIMENTO AS Dt_Venc,
    r.DTPAGAMENTO AS Dt_Pagto,
    CASE WHEN r.DTPAGAMENTO IS NULL THEN 'NAO PAGA' ELSE 'PAGA' END AS Status_Pagamento,
    snf.Status_NF_Quitada,
    snf.Parcelas_Total,
    snf.Parcelas_Pagas,
    snf.Parcelas_Abertas,
    snf.Dt_Primeiro_Venc_NF,
    snf.Dt_Ultimo_Venc_NF,
    snf.Dt_Ultimo_Pagto_NF,
    ISNULL(r.VLRDEVIDO, 0) AS Vlr_Devido,
    ISNULL(r.VLREFETIVO, 0) AS Vlr_Recebido,
    ISNULL(r.VLRDESCONTO, 0) AS Vlr_Desc,
    ISNULL(n.VLRTOTAL, 0) AS Vlr_Total_NF,

    n.NRONOTA,
    n.VEPEDIDO,

    c.CODIGO AS Cod_Cliente,
    c.RAZAO AS Cliente,
    c.CIDADE AS Cidade,
    c.UF AS UF,

    n.VENDEDOR AS Cod_Vendedor_NF,
    vn.RAZAO AS Vendedor_NF,
    ISNULL(tn.Tipo_Produto_NF, 'PECA') AS Tipo_Produto,
    ISNULL(tn.Qtd_Itens_Peca, 0) AS Qtd_Itens_Peca,
    ISNULL(tn.Qtd_Itens_Bomba, 0) AS Qtd_Itens_Bomba,
    COALESCE(vnfs.Vend_Ext_Principal, vnfs.Vend_Ext) AS Vendedor_Externo,
    COALESCE(vnfs.Vend_Int_Principal, vnfs.Vend_Int,
             CASE WHEN vn.TIPO = 'I' THEN vn.RAZAO END) AS Vendedor_Interno
FROM FN_RECEBER r
JOIN FN_NFS n ON n.CODIGO = r.NFS
LEFT JOIN FN_FORNECEDORES c ON c.CODIGO = n.CLIENTE
LEFT JOIN FN_VENDEDORES vn ON vn.CODIGO = n.VENDEDOR
LEFT JOIN VendedoresNFS vnfs ON vnfs.NFS = r.NFS
LEFT JOIN StatusNF snf ON snf.NFS = r.NFS
LEFT JOIN TipoNFS tn ON tn.NFS = r.NFS
WHERE r.NFS IS NOT NULL
    AND n.DTEMISSAO >= '{dt_ini}'
    AND n.DTEMISSAO < DATEADD(DAY, 1, '{dt_fim}')
  AND r.DTCANCELAMENTO IS NULL
  AND n.STATUSNF <> 'C'
  {filtro_ex}
ORDER BY n.DTEMISSAO, r.DTVENCIMENTO, r.NRODOCUMENTO, r.NOTASEQ
"""

    return pd.read_sql(sql, engine)


def query_controle_tipos(engine, dt_ini: str, dt_fim: str, incluir_exterior: bool) -> pd.DataFrame:
    filtro_ex = "" if incluir_exterior else "AND ISNULL(c.UF,'') <> 'EX'"

    sql = f"""
WITH TipoNFS AS (
    SELECT
        i.NFS,
        CASE
            WHEN SUM(CASE WHEN i.MATERIAL LIKE '8%' THEN 1 ELSE 0 END) > 0
             AND SUM(CASE WHEN i.MATERIAL IS NOT NULL AND i.MATERIAL NOT LIKE '8%' THEN 1 ELSE 0 END) > 0
            THEN 'AMBAS'
            WHEN SUM(CASE WHEN i.MATERIAL LIKE '8%' THEN 1 ELSE 0 END) > 0
            THEN 'BOMBA'
            WHEN SUM(CASE WHEN i.MATERIAL IS NOT NULL AND i.MATERIAL NOT LIKE '8%' THEN 1 ELSE 0 END) > 0
            THEN 'PECA'
            ELSE 'NAO_CLASSIFICADO'
        END AS Tipo_Produto
    FROM FN_NFSITENS i
    GROUP BY i.NFS
),
Base AS (
    SELECT
        r.CODIGO AS Receber,
        r.NFS,
        ISNULL(tn.Tipo_Produto, 'PECA') AS Tipo_Produto,
        ISNULL(r.VLRDEVIDO, 0) AS Vlr_Devido,
        ISNULL(r.VLREFETIVO, 0) AS Vlr_Recebido
    FROM FN_RECEBER r
    JOIN FN_NFS n ON n.CODIGO = r.NFS
    LEFT JOIN FN_FORNECEDORES c ON c.CODIGO = n.CLIENTE
    LEFT JOIN TipoNFS tn ON tn.NFS = r.NFS
    WHERE r.NFS IS NOT NULL
      AND n.DTEMISSAO >= '{dt_ini}'
      AND n.DTEMISSAO < DATEADD(DAY, 1, '{dt_fim}')
      AND r.DTCANCELAMENTO IS NULL
      AND n.STATUSNF <> 'C'
      {filtro_ex}
)
SELECT
    Tipo_Produto,
    COUNT(DISTINCT Receber) AS Titulos,
    COUNT(DISTINCT NFS) AS NFs,
    SUM(Vlr_Devido) AS Vlr_Devido,
    SUM(Vlr_Recebido) AS Vlr_Recebido
FROM Base
GROUP BY Tipo_Produto
ORDER BY Tipo_Produto
"""
    return pd.read_sql(sql, engine)


def formatar_sheet(writer, nome_aba: str, cols_data=None, cols_valor=None):
    cols_data = cols_data or []
    cols_valor = cols_valor or []

    ws = writer.book[nome_aba]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_idx = {cell.value: i + 1 for i, cell in enumerate(ws[1])}

    for col_name in cols_data:
        idx = col_idx.get(col_name)
        if not idx:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=idx).number_format = "dd/mm/yyyy"

    for col_name in cols_valor:
        idx = col_idx.get(col_name)
        if not idx:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=idx).number_format = "R$ #,##0.00"

    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, max_len + 2), 45)


def validar_consistencia(df: pd.DataFrame):
    if df.empty:
        return

    duplicados_receber = int(df["Receber"].duplicated().sum())
    if duplicados_receber > 0:
        exemplos = df[df["Receber"].duplicated(keep=False)][["Receber", "NFS", "Nro_Documento"]].head(10)
        raise ValueError(
            "Foram detectados títulos duplicados no resultado (Receber). "
            f"Qtd duplicados: {duplicados_receber}. Exemplos:\n{exemplos.to_string(index=False)}"
        )

    tipo_invalido = ~df["Tipo_Produto"].isin(["PECA", "BOMBA", "AMBAS", "NAO_CLASSIFICADO"])
    if bool(tipo_invalido.any()):
        exemplos = df.loc[tipo_invalido, ["NFS", "Receber", "Tipo_Produto"]].head(10)
        raise ValueError(
            "Foram detectados valores inválidos em Tipo_Produto. Exemplos:\n"
            f"{exemplos.to_string(index=False)}"
        )


def montar_validacao(df: pd.DataFrame, controle_sql: pd.DataFrame) -> pd.DataFrame:
    agg_df = (
        df.groupby("Tipo_Produto", dropna=False)
          .agg(
              Titulos=("Receber", "nunique"),
              NFs=("NFS", "nunique"),
              Vlr_Devido=("Vlr_Devido", "sum"),
              Vlr_Recebido=("Vlr_Recebido", "sum"),
          )
          .reset_index()
    )

    base = agg_df.merge(
        controle_sql,
        on="Tipo_Produto",
        how="outer",
        suffixes=("_Excel", "_SQL"),
    ).fillna(0)

    for col in ["Titulos_Excel", "NFs_Excel", "Titulos_SQL", "NFs_SQL"]:
        base[col] = base[col].astype(int)

    base["Dif_Titulos"] = base["Titulos_Excel"] - base["Titulos_SQL"]
    base["Dif_NFs"] = base["NFs_Excel"] - base["NFs_SQL"]
    base["Dif_Vlr_Devido"] = (base["Vlr_Devido_Excel"] - base["Vlr_Devido_SQL"]).round(2)
    base["Dif_Vlr_Recebido"] = (base["Vlr_Recebido_Excel"] - base["Vlr_Recebido_SQL"]).round(2)

    base["Status"] = base.apply(
        lambda r: "OK"
        if (
            r["Dif_Titulos"] == 0
            and r["Dif_NFs"] == 0
            and abs(r["Dif_Vlr_Devido"]) <= 0.01
            and abs(r["Dif_Vlr_Recebido"]) <= 0.01
        )
        else "DIVERGENTE",
        axis=1,
    )

    total = {
        "Tipo_Produto": "TOTAL",
        "Titulos_Excel": int(base["Titulos_Excel"].sum()),
        "Titulos_SQL": int(base["Titulos_SQL"].sum()),
        "Dif_Titulos": int(base["Dif_Titulos"].sum()),
        "NFs_Excel": int(base["NFs_Excel"].sum()),
        "NFs_SQL": int(base["NFs_SQL"].sum()),
        "Dif_NFs": int(base["Dif_NFs"].sum()),
        "Vlr_Devido_Excel": float(base["Vlr_Devido_Excel"].sum()),
        "Vlr_Devido_SQL": float(base["Vlr_Devido_SQL"].sum()),
        "Dif_Vlr_Devido": float(base["Dif_Vlr_Devido"].sum()),
        "Vlr_Recebido_Excel": float(base["Vlr_Recebido_Excel"].sum()),
        "Vlr_Recebido_SQL": float(base["Vlr_Recebido_SQL"].sum()),
        "Dif_Vlr_Recebido": float(base["Dif_Vlr_Recebido"].sum()),
        "Status": "OK"
        if (
            int(base["Dif_Titulos"].sum()) == 0
            and int(base["Dif_NFs"].sum()) == 0
            and abs(float(base["Dif_Vlr_Devido"].sum())) <= 0.01
            and abs(float(base["Dif_Vlr_Recebido"].sum())) <= 0.01
        )
        else "DIVERGENTE",
    }

    return pd.concat([base, pd.DataFrame([total])], ignore_index=True)


def gerar_excel(df: pd.DataFrame, validacao: pd.DataFrame, dt_ini: str, dt_fim: str) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo = OUTPUT / f"relatorio_506_excel_{ts}.xlsx"

    if df.empty:
        with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
            pd.DataFrame([{"Mensagem": "Sem dados no período informado."}]).to_excel(
                writer, sheet_name="Resumo", index=False
            )
        return arquivo

    cols_506 = [
        "Fl",
        "Cliente",
        "Nro_Documento",
        "Tp_Documento",
        "Dt_Emissao_NF",
        "Dt_Emissao_Titulo",
        "Dt_Venc",
        "Dt_Ultimo_Venc_NF",
        "Vlr_Devido",
        "Vlr_Total_NF",
        "Dt_Pagto",
        "Vlr_Recebido",
        "Vlr_Desc",
        "Status_Pagamento",
        "Status_NF_Quitada",
        "Parcelas_Total",
        "Parcelas_Pagas",
        "Parcelas_Abertas",
        "UF",
        "Cidade",
        "Tipo_Produto",
        "Qtd_Itens_Peca",
        "Qtd_Itens_Bomba",
        "Vendedor_Externo",
        "Vendedor_Interno",
        "Vendedor_NF",
        "NRONOTA",
        "NFS",
        "NOTASEQ",
        "Receber",
        "Cod_Cliente",
        "Cod_Vendedor_NF",
        "VEPEDIDO",
    ]
    detalhe_506 = df[cols_506].copy()

    detalhe_506 = detalhe_506.rename(columns={
        "Cliente": "Razao_Social",
        "Nro_Documento": "Nro_Documento",
        "Dt_Emissao_NF": "Dt_Emissao",
        "Dt_Venc": "Dt_Venc",
        "Vlr_Devido": "Vlr_Devido",
        "Dt_Pagto": "Dt_Pagto",
        "Vlr_Recebido": "Vlr_Recebido",
        "Vlr_Desc": "Vlr_Desc",
        "Status_Pagamento": "Status_Pagamento",
        "Vendedor_Externo": "Vendedor_Externo",
        "Vendedor_Interno": "Vendedor_Interno",
    })

    with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
        detalhe_506.to_excel(writer, sheet_name="506_Melhorado", index=False)
        validacao.to_excel(writer, sheet_name="Validacao", index=False)
        formatar_sheet(
            writer,
            "506_Melhorado",
            cols_data=["Dt_Emissao", "Dt_Emissao_Titulo", "Dt_Venc", "Dt_Ultimo_Venc_NF", "Dt_Pagto"],
            cols_valor=["Vlr_Devido", "Vlr_Total_NF", "Vlr_Recebido", "Vlr_Desc"],
        )
        formatar_sheet(
            writer,
            "Validacao",
            cols_valor=[
                "Vlr_Devido_Excel",
                "Vlr_Devido_SQL",
                "Dif_Vlr_Devido",
                "Vlr_Recebido_Excel",
                "Vlr_Recebido_SQL",
                "Dif_Vlr_Recebido",
            ],
        )

    return arquivo


def main():
    args = parse_args()

    print("Relatório 506 Excel — Contas a Receber")
    print(f"  Emissão: {args.dt_ini} até {args.dt_fim}")
    print("  Escopo geográfico: Brasil (sem EX)" if not args.incluir_exterior else "  Escopo geográfico: Brasil + exterior")

    engine = get_engine()
    df = query_506(engine, args.dt_ini, args.dt_fim, args.incluir_exterior)
    validar_consistencia(df)
    controle_sql = query_controle_tipos(engine, args.dt_ini, args.dt_fim, args.incluir_exterior)
    validacao = montar_validacao(df, controle_sql)
    arquivo = gerar_excel(df, validacao, args.dt_ini, args.dt_fim)

    print(f"\n✅ Arquivo gerado: {arquivo.resolve()}")
    if not df.empty:
        print(f"   Títulos: {df['Receber'].nunique()}")
        print(f"   NFs: {df['NFS'].nunique()}")
        print(f"   Vlr Devido total: R$ {df['Vlr_Devido'].sum():,.2f}")
        print(f"   Vlr Recebido total: R$ {df['Vlr_Recebido'].sum():,.2f}")
        print("   Distribuição Tipo_Produto:")
        print(df["Tipo_Produto"].value_counts(dropna=False).to_string())
        status_validacao = validacao.loc[validacao["Tipo_Produto"] == "TOTAL", "Status"].iloc[0]
        print(f"   Validação cruzada SQL: {status_validacao}")
    else:
        print("   Sem dados no período informado.")


if __name__ == "__main__":
    main()
