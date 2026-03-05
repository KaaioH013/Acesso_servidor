"""
fase3_margens.py — Análise de Margens e Rentabilidade (Fase 3 do Roadmap)

Gera Excel com 6 abas para o mês de referência:
  1. KPIs Resumo         — totais, cobertura, distribuição de margens
  2. Itens com Margem    — todos os itens com VLRCUSTO, margem %, alertas
  3. Por Vendedor        — margem média, faturamento, lucro bruto estimado
  4. Por Cliente         — margem média, faturamento, ranking
  5. Alertas Críticos    — margem negativa ou abaixo do threshold definido
  6. Última Compra       — itens onde última compra corrói a margem esperada

⚠️  VLRMATERIAL foi confirmado como igual ao preço de venda — NÃO é custo.
    Usa-se apenas VLRCUSTO como custo real (cobertura ~85% dos itens).

Uso:
    python fase3_margens.py                          # mês atual
    python fase3_margens.py --mes 2 --ano 2026       # fevereiro/2026
    python fase3_margens.py --mes 2 --ano 2026 --margem-critica 25
"""

import sys
import argparse
import calendar
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, "src")
from conexao import get_engine

# ── Filtros padrão peças ──────────────────────────────────────────────────────
TPVENDA_EXCLUIR = (7, 21, 5, 12, 24, 11, 26, 6, 15, 16, 8, 19, 9, 17, 53, 18, 65, 23)
TPVENDA_STR = ",".join(str(x) for x in TPVENDA_EXCLUIR)

FILTROS_PECAS = f"""
    AND p.STATUS <> 'C'
    AND i.STATUS <> 'C'
    AND i.TPVENDA NOT IN ({TPVENDA_STR})
    AND i.MATERIAL NOT LIKE '8%'
    AND i.FLAGSUB <> 'S'
    AND p.CODIGO NOT IN (
        SELECT p2.CODIGO FROM VE_PEDIDO p2
        JOIN FN_FORNECEDORES f2 ON f2.CODIGO = p2.CLIENTE
        WHERE f2.UF = 'EX'
    )
"""

OUTPUT = Path("exports")
OUTPUT.mkdir(exist_ok=True)

MARGEM_ALERTA_DEFAULT = 20.0   # % abaixo disso = alerta crítico


def parse_args():
    hoje = date.today()
    p = argparse.ArgumentParser()
    p.add_argument("--mes",  type=int, default=hoje.month)
    p.add_argument("--ano",  type=int, default=hoje.year)
    p.add_argument("--margem-critica", type=float, default=MARGEM_ALERTA_DEFAULT,
                   dest="margem_critica",
                   help="Threshold de margem crítica em %% (padrão: 20)")
    return p.parse_args()


# ── Queries ───────────────────────────────────────────────────────────────────

def query_itens_margem(engine, data_ini: str, data_fim: str) -> pd.DataFrame:
    """Todos os itens do período com colunas de custo e margem calculada."""
    sql = f"""
        SELECT
            p.CODIGO            AS Pedido,
            p.NUMINTERNO        AS Nr_Interno,
            p.DTPEDIDO          AS Dt_Pedido,
            f.RAZAO             AS Cliente,
            f.UF                AS UF,
            v.RAZAO             AS Vendedor,
            i.SEQ               AS Seq,
            i.MATERIAL          AS Material,
            i.DESCRICAO         AS Descricao,
            i.QTDE              AS Qtde,
            i.VLRUNITARIO       AS Vlr_Unit_Venda,
            i.VLRCUSTO          AS Vlr_Custo,
            i.VLRTOTAL          AS Vlr_Total_Venda,
            i.VLRDESCONTO       AS Vlr_Desconto,
            i.PERCDESCONTO      AS Perc_Desconto,
            CASE
                WHEN i.VLRCUSTO > 0 AND i.VLRUNITARIO > 0
                THEN ROUND((i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100, 2)
                ELSE NULL
            END                 AS Margem_pct,
            CASE
                WHEN i.VLRCUSTO > 0 AND i.VLRUNITARIO > 0
                THEN ROUND((i.VLRUNITARIO - i.VLRCUSTO) * i.QTDE, 2)
                ELSE NULL
            END                 AS Lucro_Bruto,
            CASE
                WHEN i.VLRCUSTO <= 0 OR i.VLRCUSTO IS NULL THEN '⚪ Sem custo cadastrado'
                WHEN (i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100 < 0   THEN '🔴 Margem negativa'
                WHEN (i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100 < 20  THEN '🟠 Crítica (< 20%%)'
                WHEN (i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100 < 40  THEN '🟡 Baixa (20-40%%)'
                WHEN (i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100 < 60  THEN '🟢 Normal (40-60%%)'
                ELSE '🔵 Alta (>= 60%%)'
            END                 AS Faixa_Margem
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p         ON p.CODIGO  = i.PEDIDO
        JOIN FN_FORNECEDORES f   ON f.CODIGO  = p.CLIENTE
        JOIN FN_VENDEDORES v     ON v.CODIGO  = p.VENDEDOR
        WHERE p.DTPEDIDO BETWEEN '{data_ini}' AND '{data_fim}'
        {FILTROS_PECAS}
        ORDER BY p.DTPEDIDO, p.CODIGO, i.SEQ
    """
    return pd.read_sql(sql, engine)


def query_por_vendedor(engine, data_ini: str, data_fim: str) -> pd.DataFrame:
    sql = f"""
        SELECT
            v.RAZAO                         AS Vendedor,
            COUNT(i.CODIGO)                 AS Itens,
            COUNT(DISTINCT p.CODIGO)        AS Pedidos,
            SUM(i.VLRTOTAL)                 AS Faturamento,
            SUM(i.VLRDESCONTO * i.QTDE)     AS Total_Desconto,
            ROUND(
                AVG(CASE WHEN i.PERCDESCONTO > 0 THEN i.PERCDESCONTO ELSE NULL END)
            , 2)                            AS Desc_Medio_pct,
            SUM(CASE WHEN i.VLRCUSTO > 0 THEN 1 ELSE 0 END)
                                            AS Itens_com_Custo,
            ROUND(AVG(
                CASE WHEN i.VLRCUSTO > 0 AND i.VLRUNITARIO > 0
                     THEN (i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100
                     ELSE NULL END
            ), 2)                           AS Margem_Media_pct,
            ROUND(AVG(
                CASE WHEN i.VLRCUSTO > 0 AND i.VLRUNITARIO > 0
                     THEN (i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100
                     ELSE NULL END
            ) - AVG(
                CASE WHEN i.PERCDESCONTO > 0 THEN i.PERCDESCONTO ELSE 0 END
            ), 2)                           AS Margem_Liquida_Est_pct,
            SUM(
                CASE WHEN i.VLRCUSTO > 0
                     THEN (i.VLRUNITARIO - i.VLRCUSTO) * i.QTDE
                     ELSE 0 END
            )                               AS Lucro_Bruto_Est
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p       ON p.CODIGO  = i.PEDIDO
        JOIN FN_VENDEDORES v   ON v.CODIGO  = p.VENDEDOR
        WHERE p.DTPEDIDO BETWEEN '{data_ini}' AND '{data_fim}'
        {FILTROS_PECAS}
        GROUP BY v.RAZAO
        ORDER BY Faturamento DESC
    """
    return pd.read_sql(sql, engine)


def query_por_cliente(engine, data_ini: str, data_fim: str) -> pd.DataFrame:
    sql = f"""
        SELECT
            f.RAZAO                         AS Cliente,
            f.CIDADE                        AS Cidade,
            f.UF                            AS UF,
            COUNT(DISTINCT p.CODIGO)        AS Pedidos,
            COUNT(i.CODIGO)                 AS Itens,
            SUM(i.VLRTOTAL)                 AS Faturamento,
            SUM(CASE WHEN i.VLRCUSTO > 0 THEN 1 ELSE 0 END)
                                            AS Itens_com_Custo,
            ROUND(AVG(
                CASE WHEN i.VLRCUSTO > 0 AND i.VLRUNITARIO > 0
                     THEN (i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100
                     ELSE NULL END
            ), 2)                           AS Margem_Media_pct,
            SUM(
                CASE WHEN i.VLRCUSTO > 0
                     THEN (i.VLRUNITARIO - i.VLRCUSTO) * i.QTDE
                     ELSE 0 END
            )                               AS Lucro_Bruto_Est
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p          ON p.CODIGO  = i.PEDIDO
        JOIN FN_FORNECEDORES f    ON f.CODIGO  = p.CLIENTE
        WHERE p.DTPEDIDO BETWEEN '{data_ini}' AND '{data_fim}'
        {FILTROS_PECAS}
        GROUP BY f.RAZAO, f.CIDADE, f.UF
        ORDER BY Faturamento DESC
    """
    return pd.read_sql(sql, engine)


def query_alertas_criticos(engine, data_ini: str, data_fim: str,
                           threshold: float) -> pd.DataFrame:
    """Itens com margem negativa ou abaixo do threshold."""
    sql = f"""
        SELECT
            p.CODIGO            AS Pedido,
            p.NUMINTERNO        AS Nr_Interno,
            p.DTPEDIDO          AS Dt_Pedido,
            f.RAZAO             AS Cliente,
            v.RAZAO             AS Vendedor,
            i.MATERIAL          AS Material,
            i.DESCRICAO         AS Descricao,
            i.QTDE              AS Qtde,
            i.VLRUNITARIO       AS Vlr_Unit_Venda,
            i.VLRCUSTO          AS Vlr_Custo,
            i.VLRTOTAL          AS Vlr_Total,
            ROUND((i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100, 2)
                                AS Margem_pct,
            CASE
                WHEN (i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100 < 0
                THEN '🔴 NEGATIVA — venda abaixo do custo'
                ELSE '🟠 CRÍTICA — abaixo de {threshold:.0f}%%'
            END                 AS Alerta,
            i.PERCDESCONTO      AS Perc_Desconto,
            i.VLRDESCONTO       AS Vlr_Desc_Unit
        FROM VE_PEDIDOITENS i
        JOIN VE_PEDIDO p         ON p.CODIGO  = i.PEDIDO
        JOIN FN_FORNECEDORES f   ON f.CODIGO  = p.CLIENTE
        JOIN FN_VENDEDORES v     ON v.CODIGO  = p.VENDEDOR
        WHERE p.DTPEDIDO BETWEEN '{data_ini}' AND '{data_fim}'
          AND i.VLRCUSTO > 0
          AND i.VLRUNITARIO > 0
          AND (i.VLRUNITARIO - i.VLRCUSTO) / i.VLRUNITARIO * 100 < {threshold}
        {FILTROS_PECAS}
        ORDER BY Margem_pct ASC
    """
    return pd.read_sql(sql, engine)


def query_ultima_compra_vs_venda(engine, data_ini: str, data_fim: str) -> pd.DataFrame:
    """
    Para cada material vendido no período, busca o preço da última compra
    e compara com o preço de venda médio.
    """
    sql = f"""
        WITH VendasPeriodo AS (
            SELECT
                i.MATERIAL,
                MAX(i.DESCRICAO)                AS Descricao,
                COUNT(*)                        AS Qtd_Vendas,
                SUM(i.QTDE)                     AS Qtde_Total,
                AVG(i.VLRUNITARIO)              AS Vlr_Venda_Medio,
                MIN(i.VLRUNITARIO)              AS Vlr_Venda_Min,
                MAX(i.VLRUNITARIO)              AS Vlr_Venda_Max,
                AVG(CASE WHEN i.VLRCUSTO > 0 THEN i.VLRCUSTO ELSE NULL END)
                                                AS Vlr_Custo_ERP
            FROM VE_PEDIDOITENS i
            JOIN VE_PEDIDO p ON p.CODIGO = i.PEDIDO
            WHERE p.DTPEDIDO BETWEEN '{data_ini}' AND '{data_fim}'
            {FILTROS_PECAS}
            GROUP BY i.MATERIAL
        ),
        UltimaCompra AS (
            SELECT
                m.MATERIAL,
                MAX(m.DATA)                     AS Dt_Ultima_Compra,
                -- Pega preço unitário da linha mais recente via subquery corr.
                (
                    SELECT TOP 1
                        ROUND(m2.VLRTOTAL / NULLIF(m2.QTDE, 0), 4)
                    FROM MT_MOVIMENTACAO m2
                    WHERE m2.MATERIAL = m.MATERIAL
                      AND m2.EVENTO = 3
                      AND m2.QTDE > 0
                      AND m2.STATUS <> 'C'
                    ORDER BY m2.DATA DESC
                )                               AS Vlr_Unit_Ultima_Compra
            FROM MT_MOVIMENTACAO m
            WHERE m.EVENTO = 3
              AND m.QTDE > 0
              AND m.STATUS <> 'C'
              AND m.MATERIAL IN (SELECT MATERIAL FROM VendasPeriodo)
            GROUP BY m.MATERIAL
        )
        SELECT
            v.MATERIAL,
            v.Descricao,
            v.Qtd_Vendas,
            v.Qtde_Total,
            ROUND(v.Vlr_Venda_Medio, 4)         AS Vlr_Venda_Medio,
            ROUND(v.Vlr_Custo_ERP, 4)           AS Vlr_Custo_ERP,
            uc.Dt_Ultima_Compra,
            ROUND(uc.Vlr_Unit_Ultima_Compra, 4) AS Vlr_Ultima_Compra,
            CASE
                WHEN uc.Vlr_Unit_Ultima_Compra IS NULL THEN NULL
                WHEN v.Vlr_Venda_Medio > 0
                THEN ROUND((v.Vlr_Venda_Medio - uc.Vlr_Unit_Ultima_Compra)
                           / v.Vlr_Venda_Medio * 100, 2)
                ELSE NULL
            END                                 AS Margem_Compra_pct,
            CASE
                WHEN uc.Vlr_Unit_Ultima_Compra IS NULL         THEN '⚪ Sem compra registrada'
                WHEN uc.Vlr_Unit_Ultima_Compra > v.Vlr_Venda_Medio THEN '🔴 Compra > Venda!'
                WHEN (v.Vlr_Venda_Medio - uc.Vlr_Unit_Ultima_Compra)
                     / v.Vlr_Venda_Medio * 100 < 20           THEN '🟠 Margem real crítica'
                WHEN (v.Vlr_Venda_Medio - uc.Vlr_Unit_Ultima_Compra)
                     / v.Vlr_Venda_Medio * 100 < 40           THEN '🟡 Margem real baixa'
                ELSE '🟢 OK'
            END                                 AS Situacao
        FROM VendasPeriodo v
        LEFT JOIN UltimaCompra uc ON uc.MATERIAL = v.MATERIAL
        ORDER BY
            CASE WHEN uc.Vlr_Unit_Ultima_Compra > v.Vlr_Venda_Medio THEN 0 ELSE 1 END,
            Margem_Compra_pct ASC
    """
    return pd.read_sql(sql, engine)


def montar_resumo(df_itens: pd.DataFrame, df_vend: pd.DataFrame,
                  df_alerta: pd.DataFrame, mes_ano: str, threshold: float) -> pd.DataFrame:
    total_fat = df_itens["Vlr_Total_Venda"].sum()
    total_itens = len(df_itens)
    com_custo = df_itens["Vlr_Custo"].gt(0).sum()
    sem_custo = total_itens - com_custo
    cob_pct = com_custo / total_itens * 100 if total_itens else 0

    margem_media = df_itens.loc[df_itens["Vlr_Custo"] > 0, "Margem_pct"].mean()
    lucro_bruto  = df_itens["Lucro_Bruto"].sum()

    dist = df_itens["Faixa_Margem"].value_counts()

    neg   = df_itens[df_itens["Margem_pct"].notna() & (df_itens["Margem_pct"] < 0)]
    crit  = df_itens[df_itens["Margem_pct"].notna() & (df_itens["Margem_pct"] >= 0) & (df_itens["Margem_pct"] < threshold)]

    linhas = [
        ("Período",                   mes_ano),
        ("Threshold margem crítica",  f"{threshold:.0f}%"),
        ("─── RESUMO GERAL ───",     ""),
        ("Faturamento total (peças)", f"R$ {total_fat:,.2f}".replace(",","X").replace(".",",").replace("X",".")),
        ("Total de itens",            f"{total_itens:,}"),
        ("Itens com VLRCUSTO",        f"{com_custo:,}  ({cob_pct:.1f}% dos itens)"),
        ("Itens sem VLRCUSTO",        f"{sem_custo:,}  — margem indisponível"),
        ("─── MARGENS ───",          ""),
        ("Margem média geral",        f"{margem_media:.1f}%"),
        ("Lucro bruto estimado",      f"R$ {lucro_bruto:,.2f}".replace(",","X").replace(".",",").replace("X",".")),
        ("─── ALERTAS ───",          ""),
        ("🔴 Margem negativa",        f"{len(neg):,} itens  |  Fat: R$ {neg['Vlr_Total_Venda'].sum():,.0f}".replace(",",".")),
        (f"🟠 Crítica (< {threshold:.0f}%)",
                                      f"{len(crit):,} itens  |  Fat: R$ {crit['Vlr_Total_Venda'].sum():,.0f}".replace(",",".")),
        ("─── DISTRIBUIÇÃO ───",     ""),
    ]
    for faixa, qtd in dist.sort_index().items():
        linhas.append((f"  {faixa}", f"{qtd:,} itens"))

    linhas.append(("─── TOP 3 VENDEDORES (MARGEM MÉDIA) ───", ""))
    top_v = df_vend.nlargest(3, "Margem_Media_pct")[["Vendedor","Margem_Media_pct","Faturamento"]]
    for _, row in top_v.iterrows():
        linhas.append((f"  {row['Vendedor'][:35]}",
                        f"{row['Margem_Media_pct']:.1f}%  |  R$ {row['Faturamento']:,.0f}".replace(",",".")))

    return pd.DataFrame(linhas, columns=["Indicador", "Valor"])


# ── Excel helpers ─────────────────────────────────────────────────────────────
CORES = {
    "🔴": "FFC7CE",
    "🟠": "FFB347",
    "🟡": "FFEB9C",
    "🟢": "C6EFCE",
    "🔵": "BDD7EE",
    "⚪": "E0E0E0",
}


def escrever_aba(writer, nome_aba: str, df: pd.DataFrame,
                 col_cor: str | None = None):
    nome = nome_aba[:31]
    df.to_excel(writer, sheet_name=nome, index=False)
    ws = writer.sheets[nome]

    hdr = PatternFill("solid", fgColor="1F4E79")
    hdr_f = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = hdr
        cell.font = hdr_f
        cell.alignment = Alignment(horizontal="center")

    if col_cor and col_cor in df.columns:
        cidx = list(df.columns).index(col_cor)
        for r in range(2, len(df) + 2):
            val = str(ws.cell(row=r, column=cidx + 1).value or "")
            hex_c = next((c for e, c in CORES.items() if val.startswith(e)), None)
            if hex_c:
                fill = PatternFill("solid", fgColor=hex_c)
                for c in range(1, len(df.columns) + 1):
                    ws.cell(row=r, column=c).fill = fill

    for i, col in enumerate(df.columns, 1):
        mx = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(i)].width = min(mx + 3, 55)

    ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    args = parse_args()
    mes, ano = args.mes, args.ano
    threshold = args.margem_critica

    ultimo_dia = calendar.monthrange(ano, mes)[1]
    data_ini = f"{ano}-{mes:02d}-01"
    data_fim = f"{ano}-{mes:02d}-{ultimo_dia:02d}"
    mes_ano  = f"{mes:02d}/{ano}"

    engine = get_engine()
    print(f"Análise de Margens — {mes_ano}  (alerta crítico: < {threshold:.0f}%)")

    print("  [1/5] Itens com margem...")
    df_itens = query_itens_margem(engine, data_ini, data_fim)

    print("  [2/5] Margem por vendedor...")
    df_vend  = query_por_vendedor(engine, data_ini, data_fim)

    print("  [3/5] Margem por cliente...")
    df_cli   = query_por_cliente(engine, data_ini, data_fim)

    print("  [4/5] Alertas de margem crítica...")
    df_alert = query_alertas_criticos(engine, data_ini, data_fim, threshold)

    print("  [5/5] Última compra vs preço de venda (pode demorar)...")
    df_comp  = query_ultima_compra_vs_venda(engine, data_ini, data_fim)

    df_resumo = montar_resumo(df_itens, df_vend, df_alert, mes_ano, threshold)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo = OUTPUT / f"margens_{mes:02d}-{ano}_{ts}.xlsx"

    print(f"\nGerando Excel: {arquivo}")
    with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
        escrever_aba(writer, "KPIs Resumo",        df_resumo)
        escrever_aba(writer, "Itens com Margem",   df_itens,  col_cor="Faixa_Margem")
        escrever_aba(writer, "Por Vendedor",        df_vend)
        escrever_aba(writer, "Por Cliente",         df_cli)
        escrever_aba(writer, "Alertas Críticos",    df_alert,  col_cor="Alerta")
        escrever_aba(writer, "Última Compra",       df_comp,   col_cor="Situacao")

    print(f"\n✅ Concluído! Arquivo: {arquivo.resolve()}")

    # ── Console ──
    com_custo = df_itens["Vlr_Custo"].gt(0).sum()
    margem_m  = df_itens.loc[df_itens["Vlr_Custo"] > 0, "Margem_pct"].mean()
    print(f"\n{'='*55}")
    print(f"  Período       : {mes_ano}")
    print(f"  Itens total   : {len(df_itens):>5,}")
    print(f"  Com VLRCUSTO  : {com_custo:>5,}  ({com_custo/len(df_itens)*100:.1f}%)")
    print(f"  Margem média  : {margem_m:>8.1f}%")
    print(f"  Lucro bruto   : R$ {df_itens['Lucro_Bruto'].sum():>10,.2f}".replace(",","X").replace(".",",").replace("X","."))
    print(f"  🔴 Alertas    : {len(df_alert):>5,} itens abaixo de {threshold:.0f}%")
    print(f"\n  Margem por vendedor:")
    for _, r in df_vend.iterrows():
        bar = "█" * int(r["Margem_Media_pct"] / 5)
        print(f"    {r['Vendedor'][:30]:<30}  {r['Margem_Media_pct']:5.1f}%  {bar}")
    print("="*55)


if __name__ == "__main__":
    main()
